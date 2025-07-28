import os
import json
from datetime import datetime
from selenium import webdriver
from axe_selenium_python import Axe
import openpyxl
from openpyxl.styles import Font

def check_wcag_compliance(url, page_name="SiteScan"):
    """
    Main function to check WCAG compliance for a given URL.
    :param url: The website URL to scan.
    :param page_name: A name for the report (e.g., "HomePage").
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    
    try:
        driver.get(url)
        axe = Axe(driver)
        axe.inject()
        results = axe.run()
        
        violations = results["violations"]
        if violations:
            print(f"Found {len(violations)} WCAG violations on {url}:")
            for v in violations:
                print(f"- Impact: {v['impact']} | Description: {v['description']}")
        else:
            print("No violations found! (Note: This is automated; manual checks still needed.)")
        
        save_reports(results, page_name)
    
    except Exception as e:
        print(f"Error during scan: {e}")
    
    finally:
        driver.quit()

def save_reports(results, page_name):
    """
    Saves JSON full results and Excel violations summary.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base_dir = "reports"
    summary_dir = os.path.join(base_dir, "summary")
    details_dir = os.path.join(base_dir, "details")
    os.makedirs(summary_dir, exist_ok=True)
    os.makedirs(details_dir, exist_ok=True)
    
    json_path = os.path.join(details_dir, f"{page_name}_{timestamp}.json")
    with open(json_path, "w") as json_file:
        json.dump(results, json_file, indent=4)
    print(f"Full JSON report saved: {json_path}")
    
    violations = results["violations"]
    if violations:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = page_name
        
        headers = ["ID", "Description", "Impact", "Help", "Help URL", "Tags"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            sheet.column_dimensions[chr(64 + col)].width = 20
        
        for row, violation in enumerate(violations, start=2):
            sheet.cell(row=row, column=1, value=violation["id"])
            sheet.cell(row=row, column=2, value=violation["description"])
            sheet.cell(row=row, column=3, value=violation["impact"])
            sheet.cell(row=row, column=4, value=violation["help"])
            sheet.cell(row=row, column=5, value=violation["helpUrl"])
            sheet.cell(row=row, column=6, value=", ".join(violation["tags"]))
        
        excel_path = os.path.join(summary_dir, f"{page_name}_{timestamp}.xlsx")
        wb.save(excel_path)
        print(f"Excel violations summary saved: {excel_path}")
    else:
        print("No violations, so no Excel summary created.")

if __name__ == "__main__":
    test_url = input("Enter the URL to scan (e.g., https://example.com): ")
    check_wcag_compliance(test_url, "AlphaScan")